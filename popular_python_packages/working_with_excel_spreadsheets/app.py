import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]

cell = sheet["a1"]
#cell = sheet.cell(row=1, column=1) # is the same code
#print(sheet.max_column) # will return a number 3
#print(sheet.max_row) # will return a number 4
# so lets print all values in the sheet
# for row in range(1, sheet.max_row + 1):
#     for col in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, col)
#         print(cell.value)

# if we want to access a range of cells, we can use this methods:
column = sheet["a1"]
cells = sheet["a1:c3"]
print(cells)
print(sheet[1:3])

# now this sheet object has a few methods you should be aware of.
# for example we use the append() method to add a row in the end of this sheet, and we have another methods also
sheet.append([1, 2, 3])
# sheet.insert_cols(0, 20)
# sheet.insert_rows(2, 10)

# and then we have to sve this workbook, and we gonna save it as a new file.
wb.save("transaction2.xlsx")